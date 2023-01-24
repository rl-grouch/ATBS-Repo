#! python
# census2010-statepop.py - dictionary created as a means of working with the excel data

import openpyxl, pprint
from openpyxl.styles import Font

print('Opening workbook...')

#open workbook
wb = openpyxl.load_workbook(r'C:\Users\steve\python\python_projects\automatestuff\spreadsheets\censuspopdata.xlsx')
sheet = wb.active

outputData = {}

#loop through workbook, saving required data in dict
for row in range(2, sheet.max_row + 1):
    
    # assign requird data in row to variables
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    # check if state exists in outputData, create new state key if not
    outputData.setdefault(state, {})

    # check if county exists in outputData, creates new county key in state if not
    outputData[state].setdefault(county, {'pop': 0})

    # adds pop integer to county key 'pop'
    outputData[state][county]['pop'] += int(pop)


# outputData = {'AK': {'Aleutians East': {'pop': 3141},
        #       'Aleutians West': {'pop': 5561},
        #       'Anchorage': {'pop': 291826}, ....

totalPop = {}

for key, value in outputData.items():
    # check / create new state key in totalPop    
    totalPop.setdefault(key, 0)
    # get pop data from output data
    for k, v in value.items():
        # print(value[k]['pop'])
        totalPop[key] += value[k]['pop']

    
    # print(f'{key} : {totalPop[key]}')


wbOutput = openpyxl.Workbook()
sheetOutput = wbOutput.active
sheetOutput.title = 'State Populations'
boldFont = Font(bold=True)
sheetOutput['A1'].value = 'State'
sheetOutput['A1'].font = boldFont
sheetOutput['B1'].value = 'Population'
sheetOutput['B1'].font = boldFont

for row in range(len(totalPop)):
    
    sheetOutput['A' + str(row+2)].value = list(totalPop.keys())[row]
    sheetOutput['B' + str(row+2)].value = list(totalPop.values())[row]

    print(f'{list(totalPop.keys())[row]} population saved to sheet.')

wbOutput.save(r'C:\Users\steve\python\python_projects\automatestuff\spreadsheets\census2010statepops.xlsx')

print('PROGRAM COMPLETE')